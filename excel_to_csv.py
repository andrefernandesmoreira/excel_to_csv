import streamlit as st
import io
import zipfile
from openpyxl import load_workbook
import datetime
from math import isclose

# --- Config da página ---
st.set_page_config(
    page_title="Andre",
    page_icon="📄",
    layout="centered",
    initial_sidebar_state="collapsed"
)

st.title("🗂️ Conversor de Excel para CSV")
st.markdown("""
Esta ferramenta converte seus **Excel (.xlsx, .xlsm)** para **CSV (UTF-8 BOM; ; \\r\\n)**
de forma consistente com o **Salvar como → CSV** do Excel, mantendo cabeçalhos, posições de colunas e quebras de linha.
""")
st.info("💡 Dica: você pode carregar vários arquivos para conversão em lote.")

uploaded_files = st.file_uploader(
    "Selecione um ou mais arquivos Excel para converter:",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True
)

def excel_like_text(cell):
    """Converte o conteúdo de uma célula para texto no estilo do Excel, imitando 'Salvar como CSV'."""
    v = cell.value
    if v is None:
        return ""

    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")

    if isinstance(v, int):
        return str(v)

    if isinstance(v, float):
        if isclose(v, round(v)):
            return str(int(round(v)))
        s = f"{abs(v):,.2f}"
        s = s.replace(",", "X").replace(".", ",").replace("X", ".")
        if v < 0:
            s = "-" + s
        return f" {s} "  # Excel coloca espaço antes e depois em valores decimais

    return str(v)

def compute_used_bounds(ws):
    """Retorna (min_col_used, max_col_used, max_row_used) do intervalo realmente usado."""
    min_col_used = None
    max_col_used = 0
    max_row_used = 0

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_has_value = False
        first_col_in_row = None
        last_col_in_row = 0

        for c_idx, v in enumerate(row, start=1):
            if v not in (None, ""):
                row_has_value = True
                if first_col_in_row is None:
                    first_col_in_row = c_idx
                last_col_in_row = c_idx

        if row_has_value:
            max_row_used = r_idx
            if min_col_used is None or first_col_in_row < min_col_used:
                min_col_used = first_col_in_row
            if last_col_in_row > max_col_used:
                max_col_used = last_col_in_row

    if min_col_used is None:
        min_col_used, max_col_used, max_row_used = 1, 1, 0

    return min_col_used, max_col_used, max_row_used

if uploaded_files:
    msg = st.empty()
    msg.info("⏳ Processando...")

    zip_buffer = io.BytesIO()
    processed = 0
    had_errors = False
    generated_csvs = []

    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zipf:
        for file in uploaded_files:
            try:
                msg.info(f"⚙️ Convertendo: **{file.name}**...")
                wb = load_workbook(file, data_only=True)
                ws = wb.active

                min_col_used, max_col_used, max_row_used = compute_used_bounds(ws)

                lines = []
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=max_row_used,
                    min_col=min_col_used,
                    max_col=max_col_used,
                    values_only=False
                ):
                    values = [excel_like_text(cell) for cell in row]
                    line = ";".join(values)
                    lines.append(line)

                txt = "\r\n".join(lines) + "\r\n"
                csv_bytes = txt.encode("utf-8-sig")

                base = file.name.rsplit(".", 1)[0]
                csv_name = f"{base}.csv"
                zipf.writestr(csv_name, csv_bytes)
                generated_csvs.append((csv_name, csv_bytes))
                processed += 1

            except Exception as e:
                st.error(f"❌ Erro ao processar **{file.name}**: {e}")
                had_errors = True

    msg.empty()

    if processed:
        if had_errors:
            st.warning(f"⚠️ Conversão concluída com erros. **{processed}** arquivo(s) convertido(s).")
        else:
            st.success(f"🎉 Conversão concluída! **{processed}** arquivo(s) convertido(s).")

        st.download_button(
            label="📥 Baixar CSVs (ZIP)",
            data=zip_buffer.getvalue(),
            file_name="arquivos_convertidos.zip",
            mime="application/zip"
        )

        # --- Validação opcional ---
        st.markdown("---")
        st.subheader("🔎 Validação (opcional): comparar dois CSVs")
        c1, c2 = st.columns(2)
        with c1:
            csv_a = st.file_uploader("CSV A", type=["csv"], key="cmp_a")
        with c2:
            csv_b = st.file_uploader("CSV B", type=["csv"], key="cmp_b")

        def normalize_csv_bytes(b: bytes) -> bytes:
            txt = b.decode("utf-8-sig")
            txt = txt.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\r\n")
            return txt.encode("utf-8")

        if csv_a and csv_b:
            a_bytes = normalize_csv_bytes(csv_a.read())
            b_bytes = normalize_csv_bytes(csv_b.read())
            if a_bytes == b_bytes:
                st.success("✅ Os dois CSVs são **exatamente iguais**.")
            else:
                st.error("❌ Os arquivos **não** são idênticos.")
                a_text = a_bytes.decode("utf-8").split("\r\n")
                b_text = b_bytes.decode("utf-8").split("\r\n")
                max_lines = max(len(a_text), len(b_text))
                for i in range(max_lines):
                    line_a = a_text[i] if i < len(a_text) else ""
                    line_b = b_text[i] if i < len(b_text) else ""
                    if line_a != line_b:
                        st.code(f"Linha {i+1}:\nA: {line_a}\nB: {line_b}")
                        break

    else:
        st.info("Nenhum arquivo processado.")

st.markdown("---")
st.markdown("👨‍💻 Desenvolvido por [Andre Fernandes Moreira]")